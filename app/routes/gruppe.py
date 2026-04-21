import datetime
from io import BytesIO

import pandas as pd
from flask import (
    Blueprint,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models import get_kompass
from app.routes.auth import require_role
from app.services.utils import heute
from app.services.gruppe_service import *

gruppe_bp = Blueprint("gruppe", __name__)


@gruppe_bp.route("/gruppen/<int:gruppe_id>", methods=["GET", "POST"])
@require_role(1)
@login_required
def gruppe(gruppe_id):
    today = datetime.date.today()
    today_name = heute()
    total_age = 0

    gruppe = get_gruppe(gruppe_id)
    mitglieder = get_mitglieder(gruppe_id)
    gruppenleiter = get_gruppenleiter(gruppe_id)

    alle_daten_mitglieder, anwesenheit = get_anwesenheit(gruppe_id, "mitglied")
    alle_daten_gruppenleiter, gruppenleiter_anwesenheit = get_anwesenheit(
        gruppe_id, "gruppenleiter"
    )
    alle_daten = list(dict.fromkeys(alle_daten_mitglieder + alle_daten_gruppenleiter))

    if today_name == gruppe["wochentag"] and str(today) not in alle_daten:
        new_day(gruppe_id=gruppe_id, datum=today)

        alle_daten_mitglieder, anwesenheit = get_anwesenheit(gruppe_id, "mitglied")
        alle_daten_gruppenleiter, gruppenleiter_anwesenheit = get_anwesenheit(
            gruppe_id, "gruppenleiter"
        )
        alle_daten = list(dict.fromkeys(alle_daten_mitglieder + alle_daten_gruppenleiter))


    for mitglied in mitglieder:
        geburtsdatum = datetime.datetime.strptime(mitglied["geburtsdatum"], "%Y-%m-%d")
        total_age += today.year - geburtsdatum.year
    avg_age = int(total_age / len(mitglieder)) if len(mitglieder) > 0 else 0

    if request.method == "POST":
        if current_user.role < 2:
            abort(403)

        update_anwesenheit(gruppe_id, "mitglied", mitglieder, request, alle_daten)
        update_anwesenheit(gruppe_id, "gruppenleiter", gruppenleiter, request, alle_daten)

        return redirect(url_for("gruppe.gruppe", gruppe_id=gruppe_id))
    
    return render_template(
        "gruppe.html",
        today_name=today_name,
        today=str(today),
        gruppe=gruppe,
        mitglieder=mitglieder,
        anwesenheit=anwesenheit,
        alle_daten=sorted(alle_daten),
        gruppenleiter=gruppenleiter,
        gruppenleiter_anwesenheit=gruppenleiter_anwesenheit,
        avg_age=avg_age,
    )


@gruppe_bp.route("/search", methods=["GET"])
@login_required
@require_role(1)
def search():
    search_query = request.args.get("query", "")
    if "mitglied" in search_query:
        search_in = "mitglieder"
        search_query = search_query.replace("mitglied", "").strip()
    else:
        search_in = "gruppenleiter"
        search_query = search_query.replace("leiter", "").strip()

    return jsonify(
        {
            "results": [
                {
                    "id": row["id"],
                    "vorname": row["vorname"],
                    "nachname": row["nachname"],
                }
                for row in search_person(search_in, search_query)
            ]
        }
    )


@gruppe_bp.route("/gruppe/<int:gruppe_id>/mitglied/<int:mitglied_id>", methods=["POST"])
@login_required
@require_role(2)
def mitglied_zu_gruppe(gruppe_id, mitglied_id):
    mitglied_zu_gruppe_hinzufuegen(gruppe_id, mitglied_id)

    return redirect(url_for("gruppe.gruppe", gruppe_id=gruppe_id))


@gruppe_bp.route(
    "/gruppenleiter_zu_gruppe/<int:gruppe_id>/<int:gruppenleiter_id>", methods=["POST"]
)
@login_required
@require_role(2)
def gruppenleiter_zu_gruppe(gruppe_id, gruppenleiter_id):
    gruppenleiter_zu_gruppe_hinzufuegen(gruppe_id, gruppenleiter_id)

    return redirect(url_for("gruppe.gruppe", gruppe_id=gruppe_id))


@gruppe_bp.route(
    "/gruppe/<int:gruppe_id>/mitglied/<int:mitglied_id>/entfernen", methods=["POST"]
)
@login_required
@require_role(3)
def mitglied_entfernen(gruppe_id, mitglied_id):
    mitglied_aus_gruppe_entfernen(gruppe_id, mitglied_id)

    return redirect(url_for("gruppe.gruppe", gruppe_id=gruppe_id))


@gruppe_bp.route(
    "/gruppe/<int:gruppe_id>/gruppenleiter/<int:gruppenleiter_id>/entfernen",
    methods=["POST"],
)
@login_required
@require_role(3)
def gruppenleiter_entfernen(gruppe_id, gruppenleiter_id):
    gruppenleiter_aus_gruppe_entfernen(gruppe_id, gruppenleiter_id)

    return redirect(url_for("gruppe.gruppe", gruppe_id=gruppe_id))


@gruppe_bp.route("/gruppe/<int:gruppe_id>/download_attendance", methods=["GET"])
@login_required
@require_role(2)
def download_attendance(gruppe_id):
    with get_kompass() as conn:
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT
                vorname,
                nachname,
                datum,
                anwesend
            FROM
                anwesenheit
                JOIN mitglieder m ON mitglied_id = m.id
            WHERE
                gruppe_id = ?
            """,
            (gruppe_id,),
        )
        anwesenheit_mitglieder = cursor.fetchall()

        cursor.execute(
            """
            SELECT
                vorname,
                nachname,
                datum,
                anwesend
            FROM
                anwesenheit_leiter
                JOIN gruppenleiter gl ON gruppenleiter_id = gl.id
            WHERE
                gruppe_id = ?
            """,
            (gruppe_id,),
        )
        anwesenheit_gruppenleiter = cursor.fetchall()

    df_m = pd.DataFrame(
        anwesenheit_mitglieder, columns=["vorname", "nachname", "datum", "anwesend"]
    )
    df_l = pd.DataFrame(
        anwesenheit_gruppenleiter, columns=["vorname", "nachname", "datum", "anwesend"]
    )

    df_m["Rolle"] = "Mitglied"
    df_l["Rolle"] = "Gruppenleiter"

    df = pd.concat([df_m, df_l])
    df["Person"] = df["vorname"] + " " + df["nachname"]
    df["datum"] = pd.to_datetime(df["datum"])

    table = pd.pivot_table(
        df, index=["Rolle", "Person"], columns="datum", values="anwesend", aggfunc="max"
    )

    table = table.fillna(0).replace({1: "X", 0: ""})
    table = table.sort_index(axis=1)
    table.columns = table.columns.strftime("%d.%m.%Y")

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Anwesenheit")

        _ = writer.book
        sheet = writer.sheets["Anwesenheit"]

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for cell in sheet["A"]:
            if cell.row != 1 and cell.value in ["Mitglied", "Gruppenleiter"]:
                cell.font = Font(bold=True)

        for row in sheet.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 3

        sheet.freeze_panes = "C2"
        sheet.auto_filter.ref = sheet.dimensions

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"anwesenheit_gruppe_{gruppe_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
