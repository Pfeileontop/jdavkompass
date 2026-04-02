import datetime
from io import BytesIO

import pandas as pd
from flask import (
    Blueprint,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models import get_kompass
from app.routes.auth import require_role
from app.utils import jugendgruppen_preview

gruppen_bp = Blueprint("gruppen", __name__)


@gruppen_bp.route("/gruppen", methods=["GET", "POST"])
@login_required
@require_role(1)
def gruppen():
    gruppen = jugendgruppen_preview()
    return render_template("gruppen.html", gruppen=gruppen)


@gruppen_bp.route("/search_mitglied", methods=["GET"])
@login_required
@require_role(1)
def search_mitglied():
    search_query = request.args.get("query", "")

    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """SELECT id, vorname, nachname FROM mitglieder WHERE vorname LIKE ? OR nachname LIKE ?""",
            ("%" + search_query + "%", "%" + search_query + "%"),
        )
        search_results_mitglieder = cursor.fetchall()

    return jsonify(
        {
            "results": [
                {
                    "id": row["id"],
                    "vorname": row["vorname"],
                    "nachname": row["nachname"],
                }
                for row in search_results_mitglieder
            ]
        }
    )


@gruppen_bp.route("/search_gruppenleiter", methods=["GET"])
@login_required
@require_role(1)
def search_gruppenleiter():
    search_query = request.args.get("query", "")

    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """SELECT id, vorname, nachname FROM gruppenleiter WHERE vorname LIKE ? OR nachname LIKE ?""",
            ("%" + search_query + "%", "%" + search_query + "%"),
        )
        search_results_gruppenleiter = cursor.fetchall()

    return jsonify(
        {
            "results": [
                {
                    "id": row["id"],
                    "vorname": row["vorname"],
                    "nachname": row["nachname"],
                }
                for row in search_results_gruppenleiter
            ]
        }
    )


@gruppen_bp.route("/gruppen/<int:gruppe_id>", methods=["GET", "POST"])
@login_required
def gruppe(gruppe_id):

    if current_user.role < 1:
        abort(403)

    if request.method == "POST" and current_user.role < 2:
        abort(403)
    today = datetime.datetime.today()
    today_name = [
        "Montag",
        "Dienstag",
        "Mittwoch",
        "Donnerstag",
        "Freitag",
        "Samstag",
        "Sonntag",
    ][today.weekday()]

    anwesenheitheute = {}
    anwesenheitshistorie = {}
    gruppenleiter_anwesenheitheute = {}
    gruppenleiter_anwesenheithistorie = {}
    alle_daten = set()
    total_age = 0
    count = 0

    with get_kompass() as conn:
        cursor = conn.cursor()

        cursor.execute(
            """SELECT id, name, wochentag, startzeit, endzeit FROM jugendgruppen WHERE id = ?""",
            (gruppe_id,),
        )
        gruppe = cursor.fetchone()
        if not gruppe:
            abort(404)

        cursor.execute(
            """SELECT g.id, g.vorname, g.nachname FROM gruppenleiter g JOIN gruppenleiter_jugendgruppen gj ON g.id = gj.gruppenleiter_id WHERE gj.jugendgruppe_id = ?""",
            (gruppe_id,),
        )
        gruppenleiter = cursor.fetchall()

        cursor.execute(
            """
            SELECT m.id, m.vorname, m.nachname, m.geburtsdatum, e.telefon 
            FROM mitglieder m
            LEFT JOIN mitglied_erziehungsberechtigte me ON m.id = me.mitglied_id
            LEFT JOIN erziehungsberechtigte e ON me.erziehungsberechtigter_id = e.id
            WHERE m.id IN (SELECT mitglied_id FROM mitglied_jugendgruppen WHERE jugendgruppe_id = ?)
        """,
            (gruppe_id,),
        )
        mitglieder = cursor.fetchall()

        for mitglied in mitglieder:
            if mitglied["geburtsdatum"]:
                geburtsdatum = datetime.datetime.strptime(
                    mitglied["geburtsdatum"], "%Y-%m-%d"
                )
                age = (
                    today.year
                    - geburtsdatum.year
                    - (
                        (today.month, today.day)
                        < (geburtsdatum.month, geburtsdatum.day)
                    )
                )
                total_age += age
                count += 1
        avg_age = int(total_age / count) if count > 0 else 0

        if gruppe["wochentag"] == today_name:
            if request.method == "POST":
                for mitglied in mitglieder:
                    anwesend = request.form.get(f"mitglied_{mitglied['id']}") == "on"
                    cursor.execute(
                        """
                        INSERT OR REPLACE INTO anwesenheit (mitglied_id, gruppe_id, datum, anwesend)
                        VALUES (?, ?, ?, ?)
                    """,
                        (mitglied["id"], gruppe_id, today.date(), anwesend),
                    )
                for gl in gruppenleiter:
                    anwesend = request.form.get(f"gruppenleiter_{gl['id']}") == "on"
                    cursor.execute(
                        """
                        INSERT OR REPLACE INTO anwesenheit_leiter (gruppenleiter_id, gruppe_id, datum, anwesend)
                        VALUES (?, ?, ?, ?)
                    """,
                        (gl["id"], gruppe_id, today.date(), anwesend),
                    )

            cursor.execute(
                """
                SELECT mitglied_id, anwesend FROM anwesenheit WHERE gruppe_id = ? AND datum = ?
            """,
                (gruppe_id, today.date()),
            )
            anwesenheitheute = {
                row["mitglied_id"]: bool(row["anwesend"]) for row in cursor.fetchall()
            }

            cursor.execute(
                """
                SELECT gruppenleiter_id, anwesend FROM anwesenheit_leiter WHERE gruppe_id = ? AND datum = ?
            """,
                (gruppe_id, today.date()),
            )
            gruppenleiter_anwesenheitheute = {
                row["gruppenleiter_id"]: bool(row["anwesend"])
                for row in cursor.fetchall()
            }

        cursor.execute(
            """
            SELECT mitglied_id, datum, anwesend FROM anwesenheit
            WHERE gruppe_id = ?
            ORDER BY datum
        """,
            (gruppe_id,),
        )
        for row in cursor.fetchall():
            mitglied_id, datum, anwesend = (
                row["mitglied_id"],
                row["datum"],
                row["anwesend"],
            )
            alle_daten.add(datum)
            if mitglied_id not in anwesenheitshistorie:
                anwesenheitshistorie[mitglied_id] = {}
            anwesenheitshistorie[mitglied_id][datum] = anwesend

        cursor.execute(
            """
            SELECT gruppenleiter_id, datum, anwesend FROM anwesenheit_leiter
            WHERE gruppe_id = ?
            ORDER BY datum
        """,
            (gruppe_id,),
        )
        for row in cursor.fetchall():
            gl_id, datum, anwesend = (
                row["gruppenleiter_id"],
                row["datum"],
                row["anwesend"],
            )
            alle_daten.add(datum)
            if gl_id not in gruppenleiter_anwesenheithistorie:
                gruppenleiter_anwesenheithistorie[gl_id] = {}
            gruppenleiter_anwesenheithistorie[gl_id][datum] = anwesend

    return render_template(
        "gruppe.html",
        today_name=today_name,
        gruppe=gruppe,
        mitglieder=mitglieder,
        anwesenheitheute=anwesenheitheute,
        anwesenheitshistorie=anwesenheitshistorie,
        alle_daten=sorted(alle_daten),
        gruppenleiter=gruppenleiter,
        gruppenleiter_anwesenheitheute=gruppenleiter_anwesenheitheute,
        gruppenleiter_anwesenheithistorie=gruppenleiter_anwesenheithistorie,
        avg_age=avg_age,
    )


@gruppen_bp.route("/gruppen/neue", methods=["POST"])
@login_required
@require_role(3)
def neue_gruppe():
    name = request.form.get("name")
    beschreibung = request.form.get("beschreibung")
    wochentag = request.form.get("wochentag")
    startzeit = request.form.get("startzeit")
    endzeit = request.form.get("endzeit")

    if not name or not wochentag or not startzeit or not endzeit:
        return "Fehlende Daten!", 400

    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO jugendgruppen (name, beschreibung, wochentag, startzeit, endzeit)
            VALUES (?, ?, ?, ?, ?)
        """,
            (name, beschreibung, wochentag, startzeit, endzeit),
        )

    return redirect(url_for("gruppen.gruppen"))


@gruppen_bp.route("/gruppen/<int:gruppe_id>/loeschen", methods=["POST"])
@login_required
@require_role(3)
def gruppe_loeschen(gruppe_id):
    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM jugendgruppen WHERE id = ?", (gruppe_id,))
        cursor.execute(
            "DELETE FROM mitglied_jugendgruppen WHERE jugendgruppe_id = ?", (gruppe_id,)
        )
        cursor.execute(
            "DELETE FROM gruppenleiter_jugendgruppen WHERE jugendgruppe_id = ?",
            (gruppe_id,),
        )

    return redirect(url_for("gruppen.gruppen"))


@gruppen_bp.route(
    "/gruppen/<int:gruppe_id>/mitglied/<int:mitglied_id>", methods=["POST"]
)
@login_required
@require_role(3)
def mitglied_zu_gruppe(gruppe_id, mitglied_id):
    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT 1 FROM mitglied_jugendgruppen 
            WHERE mitglied_id = ? AND jugendgruppe_id = ?
        """,
            (mitglied_id, gruppe_id),
        )
        existing = cursor.fetchone()

        if not existing:
            cursor.execute(
                """
                INSERT INTO mitglied_jugendgruppen (mitglied_id, jugendgruppe_id)
                VALUES (?, ?)
            """,
                (mitglied_id, gruppe_id),
            )

    return redirect(url_for("gruppen.gruppe", gruppe_id=gruppe_id))


@gruppen_bp.route(
    "/gruppenleiter_zu_gruppe/<int:gruppe_id>/<int:gruppenleiter_id>", methods=["POST"]
)
@login_required
@require_role(2)
def gruppenleiter_zu_gruppe(gruppe_id, gruppenleiter_id):
    with get_kompass() as conn:
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT 1 FROM gruppenleiter_jugendgruppen 
            WHERE gruppenleiter_id = ? AND jugendgruppe_id = ?
        """,
            (gruppenleiter_id, gruppe_id),
        )
        existing = cursor.fetchone()

        if not existing:
            cursor.execute(
                """
                INSERT INTO gruppenleiter_jugendgruppen (gruppenleiter_id, jugendgruppe_id)
                VALUES (?, ?)
            """,
                (gruppenleiter_id, gruppe_id),
            )

    return redirect(url_for("gruppen.gruppe", gruppe_id=gruppe_id))


@gruppen_bp.route(
    "/gruppen/<int:gruppe_id>/mitglied/<int:mitglied_id>/entfernen", methods=["POST"]
)
@login_required
@require_role(3)
def mitglied_entfernen(gruppe_id, mitglied_id):
    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            DELETE FROM mitglied_jugendgruppen
            WHERE jugendgruppe_id = ? AND mitglied_id = ?
        """,
            (gruppe_id, mitglied_id),
        )

    return redirect(url_for("gruppen.gruppe", gruppe_id=gruppe_id))


@gruppen_bp.route(
    "/gruppen/<int:gruppe_id>/gruppenleiter/<int:gruppenleiter_id>/entfernen",
    methods=["POST"],
)
@login_required
@require_role(3)
def gruppenleiter_entfernen(gruppe_id, gruppenleiter_id):
    with get_kompass() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            DELETE FROM gruppenleiter_jugendgruppen
            WHERE jugendgruppe_id = ? AND gruppenleiter_id = ?
        """,
            (gruppe_id, gruppenleiter_id),
        )

    return redirect(url_for("gruppen.gruppe", gruppe_id=gruppe_id))


@gruppen_bp.route("/gruppen/<int:gruppe_id>/download_attendance", methods=["GET"])
@login_required
@require_role(3)
def download_attendance(gruppe_id):
    with get_kompass() as conn:
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT vorname, nachname, datum, anwesend 
            FROM anwesenheit 
            JOIN mitglieder m ON mitglied_id = m.id 
            WHERE gruppe_id = ?
        """,
            (gruppe_id,),
        )
        anwesenheit_mitglieder = cursor.fetchall()

        cursor.execute(
            """
            SELECT vorname, nachname, datum, anwesend 
            FROM anwesenheit_leiter 
            JOIN gruppenleiter gl ON gruppenleiter_id = gl.id 
            WHERE gruppe_id = ?
        """,
            (gruppe_id,),
        )
        anwesenheit_gruppenleiter = cursor.fetchall()

    df_m = pd.DataFrame(
        anwesenheit_mitglieder, columns=["vorname", "nachname", "datum", "anwesend"]
    )
    df_l = pd.DataFrame(
        anwesenheit_gruppenleiter, columns=["vorname", "nachname", "datum", "anwesend"]
    )

    df_m["Rolle"] = "Mitglied"
    df_l["Rolle"] = "Gruppenleiter"

    df = pd.concat([df_m, df_l])
    df["Person"] = df["vorname"] + " " + df["nachname"]
    df["datum"] = pd.to_datetime(df["datum"])

    table = pd.pivot_table(
        df, index=["Rolle", "Person"], columns="datum", values="anwesend", aggfunc="max"
    )

    table = table.fillna(0).replace({1: "X", 0: ""})
    table = table.sort_index(axis=1)
    table.columns = table.columns.strftime("%d.%m.%Y")

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Anwesenheit")

        _ = writer.book
        sheet = writer.sheets["Anwesenheit"]

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for cell in sheet["A"]:
            if cell.row != 1 and cell.value in ["Mitglied", "Gruppenleiter"]:
                cell.font = Font(bold=True)

        for row in sheet.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 3

        sheet.freeze_panes = "C2"
        sheet.auto_filter.ref = sheet.dimensions

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"anwesenheit_gruppe_{gruppe_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
